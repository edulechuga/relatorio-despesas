import os
import io
import PyPDF2
from PIL import Image
import openpyxl
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from execution.logger import get_logger
from execution.db_relatorio import buscar_despesas, limpar_despesas

logger = get_logger("Gerador_Relatorio")

TEMP_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), '_recibos_pendentes')
TEMPLATE_XLSX = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'Relatorio-de-despesas-corporativas-PESSOAL.xlsx')
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'relatorios_gerados')

if not os.path.exists(TEMP_DIR):
    os.makedirs(TEMP_DIR, exist_ok=True)
if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR, exist_ok=True)

def agrupar_despesas(despesas):
    agrupado = {}
    for d in despesas:
        is_km = d.get('origem') == 'KM' or str(d.get('descricao', '')).startswith('KM')
        
        if is_km:
            chave = ("Transporte", "KM")
        else:
            chave = (d['categoria'], d['descricao'])
            
        if chave not in agrupado:
            agrupado[chave] = 0.0
        agrupado[chave] += float(d['valor'])
    return agrupado

def gerar_relatorio_excel(tipo, agrupado):
    wb = openpyxl.load_workbook(TEMPLATE_XLSX)
    ws = wb.active
    
    # 1. Muda Cabeçalho "GASTOS NO CARTÃO X"
    titulo_gastos = f"GASTOS NO CARTÃO {tipo.upper()}" if tipo.upper() == "PESSOAL" else f"GASTOS NO {tipo.upper()}"
    ws['K7'] = titulo_gastos

    # Tabela: Row 20 to 38
    # Limpar primeiro
    for r in range(20, 39):
        ws.cell(row=r, column=3).value = "" # Categoria
        ws.cell(row=r, column=5).value = "" # Data
        ws.cell(row=r, column=6).value = "" # Descricao
        ws.cell(row=r, column=9).value = "" # Obs
        ws.cell(row=r, column=10).value = "" # Reembolsável
        ws.cell(row=r, column=11).value = "" # Qtde
        ws.cell(row=r, column=12).value = "" # Preço Unit

    # Preencher dados
    data_hoje = datetime.now().strftime("%d-%b-%Y")
    
    linha_atual = 20
    for chave, soma_valor in agrupado.items():
        if linha_atual > 38:
            logger.warning("Alerta: Mais de 18 itens consolidados! Passou o limite do template.")
            break
            
        cat, desc = chave
        ws.cell(row=linha_atual, column=3).value = cat
        ws.cell(row=linha_atual, column=5).value = data_hoje
        ws.cell(row=linha_atual, column=6).value = desc
        ws.cell(row=linha_atual, column=9).value = ""
        ws.cell(row=linha_atual, column=10).value = "Sim"
        ws.cell(row=linha_atual, column=11).value = 1
        ws.cell(row=linha_atual, column=12).value = round(soma_valor, 2)
        linha_atual += 1

    # Assinatura Colaborador - Data (Row 44, Col 4)
    ws.cell(row=44, column=4).value = data_hoje

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"Relatorio_Despesas_{tipo.upper()}_{timestamp}.xlsx"
    caminho_final = os.path.join(OUTPUT_DIR, output_filename)
    wb.save(caminho_final)
    return caminho_final, output_filename

def comprimir_imagem_temporaria(img_path, qualidade=80):
    try:
        # Usa um arquivo temporário no disco p/ o reportlab ler sem crashar (evita BytesIO)
        temp_img_path = img_path + "_comp.jpg"
        with Image.open(img_path) as img:
            img = img.convert('RGB')
            # Limite máximo seguro
            max_pixels = 2500
            if img.width > max_pixels or img.height > max_pixels:
                ratio = min(max_pixels / float(img.width), max_pixels / float(img.height))
                new_w = int(img.width * ratio)
                new_h = int(img.height * ratio)
                # Resampling compatível com versões novas do Pillow (>10.0)
                try:
                    resample_filter = Image.Resampling.LANCZOS
                except AttributeError:
                    resample_filter = Image.LANCZOS
                img = img.resize((new_w, new_h), resample_filter)
            
            img.save(temp_img_path, format='JPEG', quality=qualidade, optimize=True)
            return temp_img_path
    except Exception as e:
        logger.error(f"Erro ao comprimir {img_path}: {e}")
        return None

def gerar_pdf_anexos(tipo, despesas):
    arquivos = [d['caminho_arquivo'] for d in despesas if d.get('caminho_arquivo') and os.path.exists(d['caminho_arquivo'])]
    if not arquivos:
        return None, None
        
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"Anexos_Despesas_{tipo.upper()}_{timestamp}.pdf"
    caminho_final = os.path.join(OUTPUT_DIR, output_filename)
    
    temp_pdf = os.path.join(OUTPUT_DIR, f"temp_{timestamp}.pdf")
    c = canvas.Canvas(temp_pdf, pagesize=A4)
    width, height = A4
    
    # As posições e os tamanhos devem ser ints para a nova API do ReportLab e para calcular corretamente
    posicoes = [
        (20, int(height/2 + 20), int(width/2 - 30), int(height/2 - 40)),
        (int(width/2 + 10), int(height/2 + 20), int(width/2 - 30), int(height/2 - 40)),
        (20, 20, int(width/2 - 30), int(height/2 - 40)),
        (int(width/2 + 10), 20, int(width/2 - 30), int(height/2 - 40))
    ]
    
    imagens = [f for f in arquivos if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp'))]
    pdfs = [f for f in arquivos if f.lower().endswith('.pdf')]
    
    idx = 0
    temp_images_created = []
    
    while idx < len(imagens):
        for pos in range(4):
            if idx < len(imagens):
                img_path = imagens[idx]
                try:
                    img_temp = comprimir_imagem_temporaria(img_path)
                    if img_temp:
                        temp_images_created.append(img_temp)
                        c.drawImage(img_temp, posicoes[pos][0], posicoes[pos][1], width=posicoes[pos][2], height=posicoes[pos][3], preserveAspectRatio=True, anchor='c')
                except Exception as e:
                    logger.error(f"Erro ao inserir imagem {img_path} no PDF: {e}")
                idx += 1
        c.showPage()
    c.save()
    
    # Agora mescla todos (o PDF gerado com as imagens + os PDFs originais anexados)
    merger = PyPDF2.PdfMerger()
    if len(imagens) > 0:
        merger.append(temp_pdf)
        
    for pdf_original in pdfs:
        try:
            merger.append(pdf_original)
        except Exception as e:
            logger.error(f"Ao mesclar pdf {pdf_original}: {e}")
            
    with open(caminho_final, 'wb') as f_out:
        merger.write(f_out)
        
    # Limpa
    if os.path.exists(temp_pdf):
        os.remove(temp_pdf)
        
    return caminho_final, output_filename

def limpar_relatorios_antigos(dias=7):
    import time
    try:
        agora = time.time()
        for arq in os.listdir(OUTPUT_DIR):
            caminho_arq = os.path.join(OUTPUT_DIR, arq)
            if os.path.isfile(caminho_arq):
                idade_dias = (agora - os.path.getmtime(caminho_arq)) / (24 * 3600)
                if idade_dias > dias:
                    os.remove(caminho_arq)
                    logger.info(f"Relatório antigo removido (> {dias} dias): {arq}")
    except Exception as e:
        logger.error(f"Erro ao limpar relatórios antigos: {e}")

def consolidar_geracao(tipo):
    despesas = buscar_despesas(tipo)
    if not despesas:
        return None, None
        
    agrupado = agrupar_despesas(despesas)
    
    # 1. Gera Excel
    caminho_excel, nom_e = gerar_relatorio_excel(tipo, agrupado)
    
    # 2. Gera PDF dos originais
    caminho_pdf, nom_p = gerar_pdf_anexos(tipo, despesas)
    
    # Apaga as pendentes + arq locais APENAS após sucesso
    # Primeiro apaga os arquivos
    for d in despesas:
        fp = d.get('caminho_arquivo')
        if fp and os.path.exists(fp):
            try:
                os.remove(fp)
            except:
                pass
    # Zera DB
    limpar_despesas(tipo)
    
    # Executa a limpeza da pasta de relatórios retendo apenas os últimos 7 dias
    limpar_relatorios_antigos(7)
    
    return caminho_excel, caminho_pdf
